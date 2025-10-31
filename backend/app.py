import os, io, json, base64, mimetypes, csv, sqlite3
from datetime import datetime, timezone
from pathlib import Path
from flask import Flask, request, jsonify
from flask_cors import CORS
from dotenv import load_dotenv
from openai import OpenAI
from pypdf import PdfReader
from docx import Document

# ========= إعدادات عامة =========
load_dotenv()
TEXT_MODEL   = "gpt-4.1-mini"
VISION_MODEL = "gpt-4o-mini"
OPENAI_API_KEY = os.environ.get("OPENAI_API_KEY")
client = OpenAI(api_key=OPENAI_API_KEY) if OPENAI_API_KEY else None

BASE = Path(__file__).parent
DB_PATH = BASE / "omniflow.db"

# ========= DB (sqlite3) =========
def db_conn():
    c = sqlite3.connect(DB_PATH)
    c.row_factory = sqlite3.Row
    return c

def init_db():
    with db_conn() as conn:
        conn.execute("""
          CREATE TABLE IF NOT EXISTS artifacts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            domain TEXT,
            kind TEXT,
            payload TEXT,
            created_at TEXT
          )
        """)
init_db()

def log_artifact(domain, kind, payload_obj):
    with db_conn() as conn:
        conn.execute(
            "INSERT INTO artifacts(domain,kind,payload,created_at) VALUES (?,?,?,?)",
            (domain or "", kind, json.dumps(payload_obj, ensure_ascii=False),
             datetime.now(timezone.utc).isoformat())
        )

# ========= تحميل Plugins للمجالات (اختياري) =========
DOMAIN_ANALYZERS = {}
try:
    import importlib
    PLUGINS = BASE / "plugins"
    for p in PLUGINS.glob("*.py"):
        if p.name.startswith("_"): continue
        mod = importlib.import_module(f"plugins.{p.stem}")
        if hasattr(mod, "DOMAIN_CODE") and hasattr(mod, "analyze"):
            DOMAIN_ANALYZERS[mod.DOMAIN_CODE] = mod.analyze
except Exception:
    DOMAIN_ANALYZERS = {}

# ========= Flask =========
app = Flask(__name__)
CORS(app)

@app.get("/")
def home():
    return jsonify({
        "ok": True,
        "service": "omniflow-unified",
        "endpoints": ["/api/analyze", "/api/artifacts", "/api/domains"]
    })

# ✅ إصلاح مشكلة تحميل المجالات
@app.get("/api/domains")
def domains():
    # إذا فيه Plugins حقيقية
    if DOMAIN_ANALYZERS:
        return jsonify([
            {
                "code": code,
                "name": code.capitalize(),
                "description": f"Domain: {code}"
            } for code in sorted(DOMAIN_ANALYZERS.keys())
        ])
    
    # إذا ما فيه → رجع قائمة افتراضية للواجهة
    return jsonify([
        {"code": "general", "name": "General", "description": "مساعد عام للذكاء الاصطناعي"},
        {"code": "construction", "name": "Construction", "description": "تحليل مشاريع البناء والمواد والتكلفة"},
        {"code": "medical", "name": "Medical", "description": "تحليل التقارير الطبية"},
        {"code": "cars", "name": "Cars", "description": "تشخيص أعطال السيارات وأكواد OBD"},
    ])

@app.get("/api/artifacts")
def artifacts():
    with db_conn() as conn:
        cur = conn.execute("""
          SELECT id, domain, kind, payload, created_at
          FROM artifacts
          ORDER BY id DESC
          LIMIT 200
        """)
        out = [dict(r) for r in cur.fetchall()]
    return jsonify(out)

# ========= أدوات قراءة الملفات =========
def extract_pdf(file):
    reader = PdfReader(file.stream)
    texts = []
    for p in reader.pages:
        t = p.extract_text() or ""
        if t.strip(): texts.append(t)
    return "\n\n".join(texts)

def extract_docx(file):
    data = file.read()
    doc = Document(io.BytesIO(data))
    return "\n".join([p.text for p in doc.paragraphs if p.text.strip()])

def extract_csv(file):
    txt = file.read().decode("utf-8", errors="ignore")
    return list(csv.DictReader(io.StringIO(txt)))

def extract_xlsx(file):
    try:
        import openpyxl
    except Exception:
        return {"error": "openpyxl غير متوفر على الخادم"}
    data = file.read()
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    out = {}
    for s in wb.sheetnames:
        ws = wb[s]
        rows = []
        headers = None
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                headers = [str(x) if x else "" for x in row]
                continue
            rows.append({headers[j]: (row[j]) for j in range(len(headers))})
        out[s] = rows[:200]
    return out

# ========= ذكاء نص / رؤية =========
def ai_text(msg):
    resp = client.chat.completions.create(
        model=TEXT_MODEL,
        messages=[
            {"role": "system", "content": "You are a helpful and knowledgeable AI assistant."},
            {"role": "user", "content": msg}
        ]
    )
    return resp.choices[0].message.content

def ai_vision(file, prompt):
    file.stream.seek(0)
    b64 = base64.b64encode(file.read()).decode("utf-8")
    mime = file.mimetype or "image/jpeg"
    resp = client.chat.completions.create(
        model=VISION_MODEL,
        messages=[
            {"role":"system","content":"You are a vision assistant. Describe and analyze precisely."},
            {"role":"user","content":[
                {"type":"text","text": prompt or "حلل محتوى الصورة"},
                {"type":"image_url","image_url":{"url": f"data:{mime};base64,{b64}"}}
            ]}
        ]
    )
    return resp.choices[0].message.content

# ========= Unified Endpoint =========
@app.post("/api/analyze")
def analyze():
    if client is None:
        return jsonify({"error":"OPENAI_API_KEY is missing"}), 500

    content_type = request.headers.get("Content-Type","")
    user_hint = request.form.get("prompt") or ""

    # ---- ملفات ----
    if content_type.startswith("multipart/form-data"):
        f = request.files.get("file")
        if not f:
            return jsonify({"error":"no file uploaded"}), 400

        name = f.filename or ""

        # صورة
        if f.mimetype.startswith("image/"):
            reply = ai_vision(f, user_hint)
            log_artifact("vision", "image", {"file": name, "reply": reply})
            return jsonify({"kind":"image","reply": reply})

        # PDF
        if name.lower().endswith(".pdf"):
            text = extract_pdf(f)[:8000]
            reply = ai_text(f"{user_hint or 'حلل هذا الملف'}:\n\n{text}")
            log_artifact("pdf", "doc", {"file": name, "reply": reply})
            return jsonify({"kind":"pdf","reply": reply})

        # DOCX
        if name.lower().endswith(".docx"):
            text = extract_docx(f)[:8000]
            reply = ai_text(f"{user_hint or 'حلل هذا المستند'}:\n\n{text}")
            log_artifact("docx", "doc", {"file": name, "reply": reply})
            return jsonify({"kind":"docx","reply": reply})

        # CSV
        if name.lower().endswith(".csv"):
            rows = extract_csv(f)[:200]
            reply = ai_text(f"حلل البيانات التالية:\n\n{rows}")
            log_artifact("csv", "table", {"file": name, "reply": reply})
            return jsonify({"kind":"csv","reply": reply})

        # XLSX
        if name.lower().endswith(".xlsx"):
            rows = extract_xlsx(f)
            reply = ai_text(f"حلل البيانات التالية:\n\n{json.dumps(rows)[:8000]}")
            log_artifact("xlsx", "table", {"file": name, "reply": reply})
            return jsonify({"kind":"xlsx","reply": reply})

        return jsonify({"error":"unsupported file"}), 400

    # ---- JSON / نص ----
    data = request.json or {}
    domain  = data.get("domain")
    payload = data.get("payload")

    # نص عادي
    if isinstance(payload, str):
        text = payload.strip()
        try:
            as_json = json.loads(text)
            payload = as_json
        except:
            reply = ai_text(text)
            log_artifact("chat", "message", {"q": text, "a": reply})
            return jsonify({"result": reply})

    # JSON + مجال
    if isinstance(payload, dict) and domain in DOMAIN_ANALYZERS:
        result = DOMAIN_ANALYZERS[domain](payload)
        log_artifact(domain, "analysis", result)
        return jsonify(result)

    # JSON بدون مجال
    if isinstance(payload, dict):
        reply = ai_text(f"حلل هذا JSON:\n\n{json.dumps(payload)[:8000]}")
        log_artifact("json", "summary", {"input": payload, "reply": reply})
        return jsonify({"result": reply})

    return jsonify({"error":"invalid input"}), 400


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 8000))
    app.run(host="0.0.0.0", port=port)
